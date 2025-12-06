from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Q, Sum, Subquery, OuterRef, Prefetch
from django.db.models.functions import Coalesce
from .models import Student, StudentArchive
from .forms import StudentForm, StudentSearchForm
from django.contrib.auth.decorators import login_required
from django.contrib import messages
import os
import tempfile
import subprocess
from django.http import FileResponse
from docxtpl import DocxTemplate
import re
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
import openpyxl
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse
from django.utils import timezone

@login_required
def student_list(request):
    form = StudentSearchForm(request.GET or None)
    students = Student.objects.all().order_by('full_name') 
    
    if form.is_valid():
        search = form.cleaned_data.get('search')
        institute = form.cleaned_data.get('institute')
        course = form.cleaned_data.get('course')
        dormitory = form.cleaned_data.get('dormitory')
        from_date_of_birth = form.cleaned_data.get('from_date_of_birth')
        to_date_of_birth = form.cleaned_data.get('to_date_of_birth')
        from_enrollment_year = form.cleaned_data.get('from_enrollment_year')
        to_enrollment_year = form.cleaned_data.get('to_enrollment_year')
        from_registration_date = form.cleaned_data.get('from_registration_date')
        to_registration_date = form.cleaned_data.get('to_registration_date')
        
        if search:
            students = students.filter(
                Q(full_name__icontains=search) |
                Q(phone__icontains=search)
            )
        
        if institute and institute != '---------':
            students = students.filter(institute=institute)
        
        if course and course != '---------':
            try:
                course_int = int(course)
                students = students.filter(course=course_int)
            except (ValueError, TypeError):
                pass
        
        if dormitory and dormitory != '---------':
            try:
                dormitory_int = int(dormitory)
                students = students.filter(dormitory_number=dormitory_int)
            except (ValueError, TypeError):
                pass

        # Фільтрація за датою народження
        if from_date_of_birth:
            students = students.filter(date_of_birth__gte=from_date_of_birth)
        if to_date_of_birth:
            students = students.filter(date_of_birth__lte=to_date_of_birth)

        # Фільтрація за датою вступу
        if from_enrollment_year:
            students = students.filter(enrollment_year__gte=from_enrollment_year)
        if to_enrollment_year:
            students = students.filter(enrollment_year__lte=to_enrollment_year)

        # Фільтрація за датою реєстрації
        if from_registration_date:
            students = students.filter(registration_date__gte=from_registration_date)
        if to_registration_date:
            students = students.filter(registration_date__lte=to_registration_date)
            
    paginator = Paginator(students, 50)
    page = request.GET.get('page')
    
    try:
        students_page = paginator.page(page)
    except PageNotAnInteger:
        students_page = paginator.page(1)
    except EmptyPage:
        students_page = paginator.page(paginator.num_pages)
    
    # Створюємо рядок параметрів для пагінації
    filter_params = request.GET.copy()
    if 'page' in filter_params:
        del filter_params['page']
    filter_query_string = filter_params.urlencode()
    
    context = {
        'students': students_page,
        'form': form,
        'filter_params': filter_query_string
    }
    return render(request, 'students/student_list.html', context)
    

@login_required
def student_create(request):
    if request.method == 'POST':
        form = StudentForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('student_list')
    else:
        form = StudentForm()
    
    return render(request, 'students/student_form.html', {'form': form})
@login_required
def student_update(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        form = StudentForm(request.POST, instance=student)
        if form.is_valid():
            form.save()
            # редірект назад на форму цього ж студента
            return redirect('student_update', pk=student.pk)
    else:
        form = StudentForm(instance=student)
        
    history = student.history.all().order_by('-history_date')

    diffs = []
    history_list = list(history)
    for i in range(len(history_list) - 1):
        newer = history_list[i]
        older = history_list[i + 1]
        diff = newer.diff_against(older)
        changes = [
            {
                'field': c.field,
                'old': c.old,
                'new': c.new,
            } for c in diff.changes
        ]
        if changes:
            diffs.append({
                'date': newer.history_date,
                'user': newer.history_user.username if newer.history_user else None,
                'changes': changes
            })

    return render(request, 'students/student_form.html', {
        'form': form,
        'student': student,
        'history': history,
        'diffs': diffs,
    })

@login_required
def student_delete(request, pk):
    student = get_object_or_404(Student, pk=pk)
    messages.error(request, "❌ Видалення заборонено. Будь ласка, створіть нового користувача.")
    return redirect('student_list')





@login_required
def student_contract_pdf(request, pk):
    student = get_object_or_404(Student, pk=pk)

    # Шлях до шаблону
    template_path = os.path.join(os.path.dirname(__file__), "contracts/contract.docx")

    # Завантажуємо шаблон Word
    doc = DocxTemplate(template_path)

    context = {
        "full_name": student.full_name or "____________________",
        "passport_data": student.passport_data or "____________________",
        "passport_record_number": getattr(student, "passport_record_number", "__________") or "__________",
        "institute": student.institute or "___",
        "course": student.course or "___",
        "dormitory_number": student.dormitory_number or "___",
        "dormitory_address": student.dormitory_address or "___",
        "room_number": student.room_number or "___",
        "contract_number": student.contract_number or "___",
        "contract_date": student.contract_date.strftime("%d.%m.%Y") if student.contract_date else "___",
        "termination_date": student.contract_termination_date.strftime("%d.%m.%Y") if student.contract_termination_date else "___",

        # фактична адреса
        "address": student.address or "____________________",
        "city": student.city or "____________________",
        "category": student.get_category_display() if student.category else "___",

        # телефон
        "phone": student.phone or "____________________",

        # паспорт
        "passport_issued_by": student.passport_issued_by or "____________________",
        "passport_issued_date": student.passport_issue_date.strftime("%d.%m.%Y") if student.passport_issue_date else "___",

        # домашня адреса
        "home_add_category": student.get_home_add_category_display() if student.home_add_category else "___",
        "home_add_city": student.home_add_city or "___",
        "home_add_street": student.home_add_street or "___",
        "home_add_building": student.home_add_building or "___",
        "home_add_apartment": student.home_add_apartment or "___",
    }



    # Рендеримо DOCX з даними
    doc.render(context)

    # Тимчасові файли
    tmp_dir = tempfile.mkdtemp()
    docx_path = os.path.join(tmp_dir, "contract.docx")
    pdf_path = os.path.join(tmp_dir, "contract.pdf")

    # Зберігаємо DOCX
    doc.save(docx_path)

    # Конвертація в PDF через libreoffice
    subprocess.run([
        "libreoffice", "--headless", "--convert-to", "pdf", "--outdir", tmp_dir, docx_path
    ], check=True)

    # Повертаємо PDF у браузер
    surname = student.full_name.split()[0] if student.full_name else "contract"
    
    response = FileResponse(open(pdf_path, "rb"), content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{surname}_contract.pdf"'
    return response




# Додайте цю функцію після функції student_contract_pdf
@login_required
def student_export(request):
    # Використовуємо ту саму логіку фільтрації, що і в student_list
    form = StudentSearchForm(request.GET or None)
    students = Student.objects.all().order_by('full_name')
    
    if form.is_valid():
        search = form.cleaned_data.get('search')
        institute = form.cleaned_data.get('institute')
        course = form.cleaned_data.get('course')
        dormitory = form.cleaned_data.get('dormitory')
        from_date_of_birth = form.cleaned_data.get('from_date_of_birth')
        to_date_of_birth = form.cleaned_data.get('to_date_of_birth')
        from_enrollment_year = form.cleaned_data.get('from_enrollment_year')
        to_enrollment_year = form.cleaned_data.get('to_enrollment_year')
        from_registration_date = form.cleaned_data.get('from_registration_date')
        to_registration_date = form.cleaned_data.get('to_registration_date')
        
        if search:
            students = students.filter(
                Q(full_name__icontains=search) |
                Q(phone__icontains=search)
            )
        
        if institute and institute != '---------':
            students = students.filter(institute=institute)
        
        if course and course != '---------':
            try:
                course_int = int(course)
                students = students.filter(course=course_int)
            except (ValueError, TypeError):
                pass
        
        if dormitory and dormitory != '---------':
            try:
                dormitory_int = int(dormitory)
                students = students.filter(dormitory_number=dormitory_int)
            except (ValueError, TypeError):
                pass

        # Фільтрація за датою народження
        if from_date_of_birth:
            students = students.filter(date_of_birth__gte=from_date_of_birth)
        if to_date_of_birth:
            students = students.filter(date_of_birth__lte=to_date_of_birth)

        # Фільтрація за датою вступу
        if from_enrollment_year:
            students = students.filter(enrollment_year__gte=from_enrollment_year)
        if to_enrollment_year:
            students = students.filter(enrollment_year__lte=to_enrollment_year)

        # Фільтрація за датою реєстрації
        if from_registration_date:
            students = students.filter(registration_date__gte=from_registration_date)
        if to_registration_date:
            students = students.filter(registration_date__lte=to_registration_date)

    # Створюємо нову книгу Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Студенти"

    # Додаємо заголовки
    headers = [
        'ПІБ', 'Дата народження', 'Телефон', 'ННІ', 'Курс',
        'Гуртожиток', 'Кімната', 'Дата реєстрації', 'Дата вступу', 'Дата закінчення',
        'Примітки', 'Паспортні дані', 'Дата видачі паспорта', 'Ким виданий паспорт',
        'Країна', 'Область', 'Район', 'Категорія', 'Місто', 'Адреса',
        'Домашня адреса: Країна', 'Область', 'Район', 'Категорія', 'Місто', 'Вулиця', 'Будинок', 'Квартира',
        'Дата договору', 'Номер договору', 'Дата розірвання договору',
        'Згода на реєстрацію', 'Дата реєстрації прописки', 'Гуртожиток реєстрації', 'Дата зняття з реєстрації'
    ]

    # Форматуємо заголовки
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Додаємо дані
    for row, student in enumerate(students, 2):
        ws.cell(row=row, column=1, value=student.full_name)
        ws.cell(row=row, column=2, value=student.date_of_birth.strftime("%d.%m.%Y") if student.date_of_birth else '')
        ws.cell(row=row, column=3, value=student.phone or '')
        ws.cell(row=row, column=4, value=student.institute or '')
        ws.cell(row=row, column=5, value=student.course or '')
        ws.cell(row=row, column=6, value=student.dormitory_number or '')
        ws.cell(row=row, column=7, value=student.room_number or '')
        ws.cell(row=row, column=8, value=student.registration_date.strftime("%d.%m.%Y") if student.registration_date else '')
        ws.cell(row=row, column=9, value=student.enrollment_year.strftime("%d.%m.%Y") if student.enrollment_year else '')
        ws.cell(row=row, column=10, value=student.graduation_year.strftime("%d.%m.%Y") if student.graduation_year else '')
        ws.cell(row=row, column=11, value=student.notes or '')
        ws.cell(row=row, column=12, value=student.passport_data or '')
        ws.cell(row=row, column=13, value=student.passport_issue_date.strftime("%d.%m.%Y") if student.passport_issue_date else '')
        ws.cell(row=row, column=14, value=student.passport_issued_by or '')
        ws.cell(row=row, column=15, value=student.country or '')
        ws.cell(row=row, column=16, value=student.region or '')
        ws.cell(row=row, column=17, value=student.region_rajon or '')
        ws.cell(row=row, column=18, value=student.get_category_display() if student.category else '')
        ws.cell(row=row, column=19, value=student.city or '')
        ws.cell(row=row, column=20, value=student.address or '')
        ws.cell(row=row, column=21, value=student.home_add_country or '')
        ws.cell(row=row, column=22, value=student.home_add_region or '')
        ws.cell(row=row, column=23, value=student.home_add_rajon or '')
        ws.cell(row=row, column=24, value=student.get_home_add_category_display() if student.home_add_category else '')
        ws.cell(row=row, column=25, value=student.home_add_city or '')
        ws.cell(row=row, column=26, value=student.home_add_street or '')
        ws.cell(row=row, column=27, value=student.home_add_building or '')
        ws.cell(row=row, column=28, value=student.home_add_apartment or '')
        ws.cell(row=row, column=29, value=student.contract_date.strftime("%d.%m.%Y") if student.contract_date else '')
        ws.cell(row=row, column=30, value=student.contract_number or '')
        ws.cell(row=row, column=31, value=student.contract_termination_date.strftime("%d.%m.%Y") if student.contract_termination_date else '')
        ws.cell(row=row, column=32, value='Так' if student.registration_consent else 'Ні')
        ws.cell(row=row, column=33, value=student.registration_date.strftime("%d.%m.%Y") if student.registration_date else '')
        ws.cell(row=row, column=34, value=student.registration_dormitory or '')
        ws.cell(row=row, column=35, value=student.deregistration_date.strftime("%d.%m.%Y") if student.deregistration_date else '')

    # Автоматичне налаштування ширини стовпців
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Створюємо відповідь
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"students_export_{timezone.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response




# Додаємо ці функції до вашого існуючого views.py

@login_required
def archive_student(request, pk):
    student = get_object_or_404(Student, pk=pk)
    
    # Перевіряємо, чи студент вже в архіві
    if hasattr(student, 'archive_record'):
        messages.warning(request, f'Студент {student.full_name} вже знаходиться в архіві.')
        return redirect('student_list')
    
    # Створюємо запис в архіві
    archive_record = StudentArchive(
        original_student=student,
        archived_by=request.user,
        # Копіюємо всі поля
        created_at=student.created_at,
        type=student.type,
        gender=student.gender,
        full_name=student.full_name,
        date_of_birth=student.date_of_birth,
        phone=student.phone,
        institute=student.institute,
        course=student.course,
        enrollment_year=student.enrollment_year,
        graduation_year=student.graduation_year,
        passport_data=student.passport_data,
        passport_record_number=student.passport_record_number,
        passport_issue_date=student.passport_issue_date,
        passport_issued_by=student.passport_issued_by,
        country=student.country,
        region=student.region,
        region_rajon=student.region_rajon,
        category=student.category,
        city=student.city,
        address=student.address,
        dormitory_number=student.dormitory_number,
        room_number=student.room_number,
        settlement_date=student.settlement_date,
        eviction_date=student.eviction_date,
        home_add_country=student.home_add_country,
        home_add_region=student.home_add_region,
        home_add_rajon=student.home_add_rajon,
        home_add_category=student.home_add_category,
        home_add_city=student.home_add_city,
        home_add_street=student.home_add_street,
        home_add_building=student.home_add_building,
        home_add_apartment=student.home_add_apartment,
        contract_date=student.contract_date,
        contract_number=student.contract_number,
        contract_termination_date=student.contract_termination_date,
        registration_consent=student.registration_consent,
        registration_date=student.registration_date,
        registration_dormitory=student.registration_dormitory,
        deregistration_date=student.deregistration_date,
        notes=student.notes,
    )
    archive_record.save()
    
    messages.success(request, f'Студента {student.full_name} успішно архівовано.')
    return redirect('student_list')



@login_required
def unarchive_student(request, pk):
    archive_record = get_object_or_404(StudentArchive, pk=pk)
    student_name = archive_record.full_name
    archive_record.delete()
    
    messages.success(request, f'Студента {student_name} успішно відновлено з архіву.')
    return redirect('student_archive_list')

@login_required
def student_archive_list(request):
    archives = StudentArchive.objects.all().order_by('full_name')
    
    paginator = Paginator(archives, 50)
    page = request.GET.get('page')
    
    try:
        archives_page = paginator.page(page)
    except PageNotAnInteger:
        archives_page = paginator.page(1)
    except EmptyPage:
        archives_page = paginator.page(paginator.num_pages)
    
    context = {
        'archives': archives_page,
    }
    return render(request, 'students/student_archive_list.html', context)

@login_required
def student_archive_detail(request, pk):
    archive = get_object_or_404(StudentArchive, pk=pk)
    return render(request, 'students/student_archive_detail.html', {'archive': archive})

@login_required
def combined_student_list(request):
    """Список всіх студентів (активні + архівні)"""
    active_students = Student.objects.all().order_by('full_name')
    archived_students = StudentArchive.objects.all().order_by('full_name')
    
    # Об'єднуємо в один список з позначкою типу
    all_students = []
    
    for student in active_students:
        all_students.append({
            'object': student,
            'type': 'active',
            'is_archived': False
        })
    
    for archive in archived_students:
        all_students.append({
            'object': archive,
            'type': 'archived',
            'is_archived': True
        })
    
    # Сортуємо за ПІБ
    all_students.sort(key=lambda x: x['object'].full_name)
    
    paginator = Paginator(all_students, 50)
    page = request.GET.get('page')
    
    try:
        students_page = paginator.page(page)
    except PageNotAnInteger:
        students_page = paginator.page(1)
    except EmptyPage:
        students_page = paginator.page(paginator.num_pages)
    
    context = {
        'students': students_page,
        'show_archived': True
    }
    return render(request, 'students/student_list.html', context)





from .models import Penalty
from .forms import PenaltyForm, PenaltySearchForm, PenaltyCancellationForm
from django.db.models import Q, Sum

from django.db.models import Q, Sum

from django.db.models import Q, Sum, Subquery, OuterRef
from django.db.models.functions import Coalesce

from django.db.models import Prefetch

@login_required
def penalty_list(request):
    form = PenaltySearchForm(request.GET or None)
    
    # Створюємо оптимізований запит з prefetch
    penalties = Penalty.objects.all().select_related('student', 'created_by').prefetch_related(
        Prefetch(
            'student__penalties',
            queryset=Penalty.objects.filter(status='active'),
            to_attr='active_penalties_cache'
        ),
        Prefetch(
            'student__penalty_reductions',
            to_attr='penalty_reductions_cache'
        ),
        Prefetch(
            'reductions',
            to_attr='penalty_reductions_for_this'
        )
    )
    
    # Підзапит для отримання загальної суми активних штрафів кожного студента (без відпрацювань)
    total_points_subquery = Penalty.objects.filter(
        student_id=OuterRef('student_id'),
        status='active'
    ).values('student_id').annotate(
        total=Sum('points')
    ).values('total')
    
    # Додаємо анотацію з загальною сумою балів студента без відпрацювань
    penalties = penalties.annotate(
        student_total_points=Coalesce(Subquery(total_points_subquery), 0)
    )
    
    if form.is_valid():
        student_search = form.cleaned_data.get('student_search')
        dormitory = form.cleaned_data.get('dormitory')
        status = form.cleaned_data.get('status')
        severity = form.cleaned_data.get('severity')
        date_from = form.cleaned_data.get('date_from')
        date_to = form.cleaned_data.get('date_to')
        institute = form.cleaned_data.get('institute')
        min_points = form.cleaned_data.get('min_points')
        
        if student_search:
            penalties = penalties.filter(
                Q(student__full_name__icontains=student_search)
            )
        
        if dormitory:
            penalties = penalties.filter(student__dormitory_number=dormitory)
            
        if status == 'active':
            penalties = penalties.filter(status='active')
            
        if severity:
            penalties = penalties.filter(severity=severity)
            
        if date_from:
            penalties = penalties.filter(penalty_date__gte=date_from)
            
        if date_to:
            penalties = penalties.filter(penalty_date__lte=date_to)
        
        if institute:
            penalties = penalties.filter(student__institute=institute)
        
        # ФІЛЬТР ЗА СУМОЮ БАЛІВ (без відпрацювань)
        if min_points:
            try:
                min_points_int = int(min_points)
                penalties = penalties.filter(student_total_points__gt=min_points_int)
            except ValueError:
                pass
    
    # Сортування спочатку за загальною сумою балів (більші першими), потім за датою
    penalties = penalties.order_by('-student_total_points', '-penalty_date', '-created_at')
    
    # Пагінація
    paginator = Paginator(penalties, 50)
    page = request.GET.get('page')
    
    try:
        penalties_page = paginator.page(page)
    except PageNotAnInteger:
        penalties_page = paginator.page(1)
    except EmptyPage:
        penalties_page = paginator.page(paginator.num_pages)
    
    # Обчислюємо загальну суму балів для кожного студента З урахуванням відпрацювань
    for penalty in penalties_page:
        # Для відображення використовуємо властивість, яка враховує відпрацювання
        penalty.display_total = penalty.student.total_penalty_points_with_reductions
        
        # Обчислюємо відпрацювання для цього конкретного штрафу
        if hasattr(penalty, 'penalty_reductions_for_this'):
            # Використовуємо кешовані дані
            penalty.total_reductions_for_this_penalty = sum(
                r.points_reduced for r in penalty.penalty_reductions_for_this
            )
        else:
            # Якщо немає кешу, обчислюємо через aggregate
            penalty.total_reductions_for_this_penalty = penalty.reduced_points
    
    context = {
        'penalties': penalties_page,
        'form': form,
        'active_tab': 'penalties'
    }
    return render(request, 'students/penalty_list.html', context)

@login_required
def penalty_create(request):
    if request.method == 'POST':
        form = PenaltyForm(request.POST)
        if form.is_valid():
            penalty = form.save(commit=False)
            penalty.created_by = request.user
            penalty.save()
            messages.success(request, f'Штрафний бал успішно додано для {penalty.student.full_name}')
            return redirect('penalty_list')
    else:
        form = PenaltyForm()
    
    context = {
        'form': form,
        'title': 'Додати штрафний бал'
    }
    return render(request, 'students/penalty_form.html', context)

@login_required
def penalty_create_for_student(request, student_id):
    student = get_object_or_404(Student, pk=student_id)
    
    if request.method == 'POST':
        form = PenaltyForm(request.POST)
        if form.is_valid():
            penalty = form.save(commit=False)
            penalty.created_by = request.user
            penalty.save()
            messages.success(request, f'Штрафний бал успішно додано для {student.full_name}')
            return redirect('student_update', pk=student.pk)
    else:
        form = PenaltyForm(initial={'student': student})
    
    context = {
        'form': form,
        'student': student,
        'title': f'Додати штрафний бал для {student.full_name}'
    }
    return render(request, 'students/penalty_form.html', context)

@login_required
def penalty_cancel(request, pk):
    penalty = get_object_or_404(Penalty, pk=pk)
    
    if penalty.status != 'active':
        messages.warning(request, 'Цей штрафний бал вже скасовано.')
        return redirect('penalty_list')
    
    if request.method == 'POST':
        form = PenaltyCancellationForm(request.POST, instance=penalty)
        if form.is_valid():
            penalty = form.save(commit=False)
            penalty.status = 'cancelled'
            penalty.cancelled_by = request.user
            penalty.cancelled_at = timezone.now()
            penalty.save()
            
            messages.success(request, f'Штрафний бал успішно скасовано для {penalty.student.full_name}')
            return redirect('penalty_list')
    else:
        form = PenaltyCancellationForm(instance=penalty)
    
    context = {
        'form': form,
        'penalty': penalty,
        'title': f'Скасувати штрафний бал для {penalty.student.full_name}'
    }
    return render(request, 'students/penalty_cancel.html', context)

@login_required
def students_with_penalties(request):
    """Сторінка зі списком студентів, які мають штрафні бали"""
    students_with_penalties = Student.objects.filter(
        penalties__status='active'
    ).annotate(
        total_points=Sum('penalties__points')
    ).filter(
        total_points__gt=0
    ).distinct().order_by('-total_points', 'full_name')
    
    # Форма пошуку
    form = PenaltySearchForm(request.GET or None)
    
    if form.is_valid():
        student_search = form.cleaned_data.get('student_search')
        dormitory = form.cleaned_data.get('dormitory')
        
        if student_search:
            students_with_penalties = students_with_penalties.filter(
                full_name__icontains=student_search
            )
        
        if dormitory:
            students_with_penalties = students_with_penalties.filter(dormitory_number=dormitory)
    
    paginator = Paginator(students_with_penalties, 50)
    page = request.GET.get('page')
    
    try:
        students_page = paginator.page(page)
    except PageNotAnInteger:
        students_page = paginator.page(1)
    except EmptyPage:
        students_page = paginator.page(paginator.num_pages)
    
    context = {
        'students': students_page,
        'form': form,
        'active_tab': 'students_with_penalties',
        'title': 'Студенти з штрафними балами'
    }
    return render(request, 'students/students_with_penalties.html', context)


@login_required
def penalty_delete(request, pk):
    penalty = get_object_or_404(Penalty, pk=pk)
    
    if request.method == 'POST':
        student_name = penalty.student.full_name
        penalty.delete()
        messages.success(request, f'Штрафний бал для {student_name} успішно видалено')
        return redirect('penalty_list')
    
    context = {
        'penalty': penalty,
    }
    return render(request, 'students/penalty_delete.html', context)




from .models import PenaltyReduction
from .forms import PenaltyReductionForm

@login_required
def penalty_reduction_create(request, student_id=None, penalty_id=None):
    """Створення відпрацювання для студента або конкретного штрафу"""
    student = None
    penalty = None
    
    if student_id:
        student = get_object_or_404(Student, pk=student_id)
    
    if penalty_id:
        penalty = get_object_or_404(Penalty, pk=penalty_id)
        student = penalty.student  # Студент береться зі штрафу
    
    if request.method == 'POST':
        form = PenaltyReductionForm(request.POST, student_id=student.id if student else None)
        if form.is_valid():
            reduction = form.save(commit=False)
            reduction.created_by = request.user
            
            # Якщо передано penalty_id, автоматично встановлюємо зв'язок
            if penalty:
                reduction.penalty = penalty
            
            reduction.save()
            
            messages.success(request, 
                f'Відпрацювання {reduction.points_reduced} балів успішно зараховано для {reduction.student.full_name}'
            )
            
            # Повертаємося на ту сторінку, звідки прийшли
            if penalty_id:
                return redirect('penalty_list')
            elif student_id:
                return redirect('student_update', pk=student_id)
            else:
                return redirect('penalty_list')
    else:
        initial_data = {}
        if student:
            initial_data['student'] = student
        if penalty:
            initial_data['penalty'] = penalty
            
        form = PenaltyReductionForm(initial=initial_data, student_id=student.id if student else None)
    
    # Обчислюємо доступні для списання бали
    available_points = 0
    if student:
        total_penalties = student.penalties.filter(status='active').aggregate(
            total=Sum('points')
        )['total'] or 0
        
        total_reductions = student.penalty_reductions.aggregate(
            total=Sum('points_reduced')
        )['total'] or 0
        
        available_points = max(0, total_penalties - total_reductions)
    
    context = {
        'form': form,
        'student': student,
        'penalty': penalty,
        'available_points': available_points,
        'title': 'Додати відпрацювання штрафних балів'
    }
    return render(request, 'students/penalty_reduction_form.html', context)
from django import forms

@login_required
def penalty_reduction_list(request):
    """Список всіх відпрацювань"""
    reductions = PenaltyReduction.objects.all().select_related(
        'student', 'penalty', 'created_by'
    ).order_by('-reduction_date', '-created_at')
    
    # Форма фільтрації
    class ReductionSearchForm(forms.Form):
        student_search = forms.CharField(
            required=False, 
            label='Пошук студента',
            widget=forms.TextInput(attrs={'placeholder': 'ПІБ студента'})
        )
        date_from = forms.DateField(
            required=False,
            label='Дата від',
            widget=forms.DateInput(attrs={'type': 'date'})
        )
        date_to = forms.DateField(
            required=False,
            label='Дата до',
            widget=forms.DateInput(attrs={'type': 'date'})
        )
    
    form = ReductionSearchForm(request.GET or None)
    
    if form.is_valid():
        student_search = form.cleaned_data.get('student_search')
        date_from = form.cleaned_data.get('date_from')
        date_to = form.cleaned_data.get('date_to')
        
        if student_search:
            reductions = reductions.filter(
                Q(student__full_name__icontains=student_search)
            )
        
        if date_from:
            reductions = reductions.filter(reduction_date__gte=date_from)
        
        if date_to:
            reductions = reductions.filter(reduction_date__lte=date_to)
    
    paginator = Paginator(reductions, 50)
    page = request.GET.get('page')
    
    try:
        reductions_page = paginator.page(page)
    except PageNotAnInteger:
        reductions_page = paginator.page(1)
    except EmptyPage:
        reductions_page = paginator.page(paginator.num_pages)
    
    context = {
        'reductions': reductions_page,
        'form': form,
        'title': 'Відпрацювання штрафних балів'
    }
    return render(request, 'students/penalty_reduction_list.html', context)

@login_required
def penalty_reduction_delete(request, pk):
    """Видалення відпрацювання"""
    reduction = get_object_or_404(PenaltyReduction, pk=pk)
    
    if request.method == 'POST':
        student_name = reduction.student.full_name
        points = reduction.points_reduced
        reduction.delete()
        messages.warning(request, f'Відпрацювання {points} балів для {student_name} успішно видалено')
        return redirect('penalty_reduction_list')
    
    context = {
        'reduction': reduction,
    }
    return render(request, 'students/penalty_reduction_delete.html', context)